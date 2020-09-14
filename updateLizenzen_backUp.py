import numpy as np
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color, Font, Border, colors
from random import randint
from contextlib import redirect_stdout

#___________________________________ LOGGING _______________________________
''' Leitet Konsolenausgabe in txtdatei um '''

def createTimeStamp():
    secondsSinceEpoch = time.time()
    timeObj = time.localtime(secondsSinceEpoch)
    timestamp = '%d%d%d_%d%d%d' % (timeObj.tm_year, timeObj.tm_mon, timeObj.tm_mday, timeObj.tm_hour, timeObj.tm_min, timeObj.tm_sec)
    return timestamp

dateiname = 'lizenzenUpdateLog_'+createTimeStamp()+'.txt'

def print_(text, logdatei = dateiname):
    print(text)
    with open(logdatei, 'a') as f:
        with redirect_stdout(f):
            print(text)
#___________________________________________________________________________


EXCEL_DATEI = 'daten.xlsx'                  # Pfad der Exceldatei
EXCEL_TABELLE_ALT = 'lizenzpool_basis'      # Tabellenname der Basisdaten
EXCEL_TABELLE_NEU = 'lizenzpool_neu'        # Tabellenname der Neuen Daten

lizenzpool_alt = pd.read_excel(open(EXCEL_DATEI, 'rb'),sheet_name=EXCEL_TABELLE_ALT)
lizenzpool_neu = pd.read_excel(open(EXCEL_DATEI, 'rb'),sheet_name=EXCEL_TABELLE_NEU)

def getValueChangeIndices(dataframe, spalte):
    ''' Diese Funktion ermittelt die Zeilen bei denen sich Werte
    verändern und gibt Spaltenindizes und Spaltenwert zurück {2 : "komsa"; 7 : "ecs";...} '''

    ranges = []
    changes = []
    values = {}

    row_count = dataframe.shape[0]

    for i in range(row_count-1):
        if dataframe.loc[i][spalte] == dataframe.loc[i+1][spalte]:
            ranges.append(i)
        else:
            ranges.append(i)
            ranges.append("-")

    changes.append(0)
    for i in range(len(ranges)):
        if ranges[i] == "-":
            changes.append(ranges[i+1])

    for i in range(len(changes)):
        values.update({changes[i]:dataframe.loc[changes[i]][spalte]})

    return values

def updateList(neu=lizenzpool_neu,alt=lizenzpool_alt, loescheDoppelteLizenzid = True):
    ''' Aktualisiert die Anzahlwerte der alten Liste durch die Anzahlwerte der neuen Liste anhand
        des Primärschlüssels "lizenzid" und fügt die Reihen ohne doppelte Lizenzids der neuen Liste hinzu'''

    print_('Beginne Listenupdate um: '+str(createTimeStamp())+".")
    print_("Listenlaenge des neuen Lizenzpools: " + str(len(neu)))

    # berechnet anhand der Schnittmenge der LizenzIds aus altem und neuem Dokument doppelte Lizenzids
    doppleteLizenzIDs = set(neu['lizenzid']).intersection(set(alt['lizenzid']))

    #einfacher Count um Differenz der Anzahlwerte der Reihen mit gleichen Lizenzids zu zählen
    anzahlDifferenz_cnt = 1

    for lizenzid in doppleteLizenzIDs:
        ''' Jede Iteration des For Loops befasst sich mit einer doppelt vorkommenden Lizenzid (in beiden Dokumenten) '''

        # Anzahlwerte der dopplet Vorkommenden Lizenzids des neuen Dokuments
        neueAnzahlWerte = neu.loc[neu['lizenzid'] == lizenzid, 'anzahl']
        neueAnzahlWerte = int(neueAnzahlWerte.tolist()[0])

        # Anzahlwerte der doppelt vorkommenden Lizenzids des alten Dokuments
        alteAnzahlWerte = alt.loc[alt['lizenzid'] == lizenzid, 'anzahl']
        alteAnzahlWerte = int(alteAnzahlWerte.tolist()[0]) # Kovertierung in integer-Werte

        # berechnet Zeilenindex [43,567,865...] der Zeilen bei der Zeile des neuen Dokuments (an der Stelle wo die doppelte Lizenzid der Iteration ist)
        zeilenIndex_neu = neu.loc[neu['lizenzid'] == lizenzid].index[0]

        if alteAnzahlWerte != neueAnzahlWerte:
            ''' Falls die Anzahlwerte im alten und neuen Dokument sich unterscheiden '''

            print_('Anzahldifferenz '+str(anzahlDifferenz_cnt)+ ' - bei Lizenzid: "' + str(lizenzid) + '" Alt: ' + str(alteAnzahlWerte) + " / " + 'Neu: '+str(neueAnzahlWerte) + '.')

            # setzt die Anzahl des alten Dokuments bei identischer Lizenzid auf die Anzahl des neuen Dokuments
            zeilenIndex_alt = alt.loc[alt['lizenzid'] == lizenzid].index[0] # ermittlet den Zeilenindex des alten Dokuments bei dem Lizenzid der Iteration vorkommt
            alt.at[zeilenIndex_alt, "anzahl"] = neueAnzahlWerte

            print_("Zeile: '"+str(zeilenIndex_neu)+"' Spalte 'anzahl' altenWert: '"+str(alteAnzahlWerte)+" durch neue '"+str(neueAnzahlWerte) +"' ersetzt.")

            # entfernt die Zeile aus der neuen Datei
            neu.drop(zeilenIndex_neu, axis=0, inplace=True)

            anzahlDifferenz_cnt += 1    # erhöht Differenz_count

        else:
            ''' Falls die Anzahlwerte der Zeilen mit identischen Lizenzids gleich sind können alten Zeilen erhalten bleiben oder gelöscht werden'''

            if loescheDoppelteLizenzid: # True = doppelte Werte werden gelöscht / False = doppelte Werte bleiben erhalten
                print_('Doppelte Lizenzid: "'+str(lizenzid)+ '" - loesche Zeile: "'+str(zeilenIndex_neu)+ '" aus neuem Dokument...')
                neu.drop(index=[zeilenIndex_neu],inplace=True)

    # Statusausgaben
    print_(str(len(doppleteLizenzIDs)+1) + ' Vorkommen doppelter Lizenzids')
    print_("Listenlaenge des neuen Lizenzpools nach Bearbeitung: " + str(len(neu)))

    # verbindet alte und neue Liste
    df1_concatenated = pd.concat([alt,neu])
    # kovertiert DataFrame in eine Liste zur weiteren Verarbeitung
    basisdaten = df1_concatenated.values.tolist()
    return basisdaten   # returned die Liste

def listToExcel(dateiname, liste):
    ''' Kovertiert Python Liste erst in Pandas Dataframe, sortiert es nach Typ und Leistungsmerkmal und speichert es dann in einer Excel
        Rückgabewert ist das End-Dataframe '''

    df1 = pd.DataFrame(liste,columns=['typ','leistungsmerkmal','lac','erstelldatum','firmenname','sid','lizenzid','sachnr','komsa','anzahl','einkauf',
                                            'hek','anzahl_nach_entnahme','anzahl_entnommen','auftragsnr','entnommen_datum','kunde_neu','entnommen_durch','sid_neu'])
    df1_sorted = df1.sort_values(['typ','leistungsmerkmal'], ascending=[True,True])
    print_('Liste nach "Typ" und "Leistungsmerkmal" sortiert')

    try:
        ''' Error-Handling ändert Dateiname, falls Dateiname bereits vorhanden ist '''

        dateiname = dateiname + ".xlsx"
        df1_sorted.to_excel(dateiname)
        print_("Liste in Exceldatei: " + str(dateiname) + ' gespeichert.')
    except PermissionError:
        dateiname = dateiname + str(randint(1,500)) + '.xlsx'
        df1_sorted.to_excel(dateiname)
        print_("Liste in Exceldatei: '" + str(dateiname) + "' gespeichert.")

def loescheExcelSpalte(dateipfad, excelSpalte):
    ''' Diese Funktion lädt die erste Tabelle einer Excel Datei und löscht eine komplette Spalte '''

    wb = load_workbook(dateipfad)             # wb = workbook
    ws = wb.active                            # ws = worksheet
    ws.delete_cols(excelSpalte)
    wb.save(dateipfad)
    print_("Spalte: '" + str(excelSpalte) + "' aus '" +str(dateipfad)+ "' geloescht.")

def mergeCells(dateipfad, spalte):
    ''' Diese Funktion verbindet Zellen einer Spalte die die gleichen Werte enthält automatisch'''

    # lädt Excel Tabelle ins Dataframe
    dataframe = pd.read_excel(dateipfad)
    # Holt sich die Zeilennummer an denen sich die Werte der Spaltenzelle ändern
    valueChangeList = [zahl for zahl in getValueChangeIndices(dataframe,spalte)]
    # wandelt die ExcelSpaltenZahl in Buchstaben um das korrekte Excelformat einzubehalten
    excelSpalte = nummerZuBuchstabe(dataframe.columns.get_loc(spalte))

    wb = load_workbook(dateipfad)   # wb = workbook
    ws = wb.active                  # ws = worksheet

    for i in range(len(valueChangeList)):
        ''' Jede Iteration geht Anfangs und Endzeile einer Excelzellen-Wertänderung durch '''

        # Für den ersten Wert
        if i == 0:
            startZeile = 2  # Pythonliste fängt bei 0 an Exceltabelle bei 2 (Annahme, 1. Zeile Erklärzeile)
            endZeile = int(valueChangeList[i+1]) + 1
        else: # Für alle anderen Werte
            if i < len(valueChangeList) - 1:
                # Alle Fälle außer der letzte
                startZeile =int(valueChangeList[i]) + 2
                endZeile = int(valueChangeList[i+1]) + 1

        # Erstellt Namen der Start und Endzellen bei denen keine Werteveränderung vorliegt
        # (also die die gemerged werden können z.B. 'A3:A15')
        startZelle = excelSpalte + str(startZeile)
        endZelle = excelSpalte + str(endZeile)

        # openpyxl Merge-Befehl
        ws.merge_cells(startZelle +':'+endZelle)

    print_('Zellen in Spalte: "'+str(spalte)+'" korrekt gemerged.')
    wb.save(dateipfad)

def nummerZuBuchstabe(number):
    EXCEL_SPALTEN = ['-',
                    'A','B','C','D','E','F','G','H','I','J',
                    'K','L','M','N','O','P','Q','R','S','T',
                    'U','V','W','X','Y','Z','AA','AB','AC','AD',
                    'AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN',
                    'AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX',
                    'AY','AZ','BA','BB','BC']
    return EXCEL_SPALTEN[number + 1]

def autoSpaltenBreite(dateipfad):
    wb = load_workbook(dateipfad)             # wb = workbook
    ws = wb.active                            # ws = worksheet

    # Funktionsdefinition zur Ermittlung der Textlaenge
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    wb.save(dateipfad)

def updateAnzahl(dateipfad):
    df1_sorted = pd.read_excel(dateipfad)
    print_(df1_sorted.duplicated(subset=['lizenzid']))
    #df1_sorted.to_excel(dateipfad)

def duplikateEntfernen(dateipfad, spalte):
    ''' Entfernt alle Zeilen bei denen Dupilkate einer Spalte auftauchen '''

    df1_sorted = pd.read_excel(dateipfad) # liest Excel-Datei ein

    # drop_duplicates entfernt die Duplikate einer Spalte
    # keep=first: erstes Vorkommen wird behalten
    # keep=last: letztes Vorkommen wird behalten
    # keep=False: Kein Wert wird behalten
    df1_sorted.drop_duplicates(subset=spalte, keep='first',inplace=True)
    df1_sorted.to_excel(dateipfad)
    print_("Duplikate der Spalte: '" + str(spalte) + "' der Excel-Datei: '"+str(dateipfad)+"' entfernt")

def spalteFaerben(dateipfad, zelle,farbe):
    ''' Diese Funktion kann eine einzelne Zelle einer Exceldatei färben '''

    def color(farbe):
        # Switch-Statement dass Farben anhand eines Strings zurück gibt
        switcher={
                'white':PatternFill(start_color='00FFFFFF',end_color='00FFFFFF',fill_type='solid'),
                'grey1':PatternFill(start_color='00C0C0C0',end_color='00C0C0C0',fill_type='solid'),
                'grey2':PatternFill(start_color='00969696',end_color='00969696',fill_type='solid'),
                'grey3':PatternFill(start_color='00808080',end_color='00808080',fill_type='solid'),
                'green':PatternFill(start_color='0099CC00',end_color='0099CC00',fill_type='solid'),
                'red':PatternFill(start_color='00FF0000',end_color='00FF0000',fill_type='solid'),
                'yellow':PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid'),
                'orange':PatternFill(start_color='00FF9900',end_color='00FF9900',fill_type='solid'),
                'blue':PatternFill(start_color='0033CCCC',end_color='0033CCCC',fill_type='solid'),
                'purple':PatternFill(start_color='00CC99FF',end_color='00CC99FF',fill_type='solid'),
             }
        return switcher.get(farbe,"Farbe nicht vorhanden")

    wb = load_workbook(dateipfad)               # wb = workbook
    tabelle = wb.active                         # ws = worksheet
    tabelle[zelle].fill = color(farbe)          # Fülle Zelle mit Farbe aus color-Funktion
    print_('Zelle: "' + str(zelle) + '" der Datei "'+str(dateipfad)+'" in Farbe "' +str(farbe)+'" umgefaerbt.')
    wb.save(dateipfad)

def ergebnisKopfzeileDesign(dateipfad = 'ergebnis.xlsx'):
    for i in range(0,9):    # Spalte 9 wird ausgeschlossen
        # iteriert von A1 bis H1
        zelle = str(nummerZuBuchstabe(i)) + '1'
        spalteFaerben(dateipfad,zelle,'yellow')
    spalteFaerben(dateipfad,"I1","green")
    spalteFaerben(dateipfad,"J1","yellow")
    for i in range(10,19):
        # iteriert von K1 bis S1
        zelle = str(nummerZuBuchstabe(i)) + '1'
        spalteFaerben(dateipfad,zelle,'green')
    for spalte in range(1,20):
        textZentrieren(dateipfad,1,spalte)

def zeilenHoehe(dateipfad='ergebnis.xlsx', zeile=1, hoehe=40):
    wb = load_workbook(dateipfad)
    ws = wb.active
    ws.row_dimensions[zeile].height = hoehe
    print_("Zeilenhoehe angepasst: "+str(hoehe)+ "cm der Zeile: '" +str(zeile)+ "' des Dokuments '" +str(dateipfad)+"'.")
    wb.save(dateipfad)

def textZentrieren(dateipfad, zeile, spalte):
    wb = load_workbook(dateipfad)
    ws = wb.active
    ws.cell(zeile, spalte).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    print_("Text von Zelle:("+str(zeile)+","+str(spalte)+") zentriert.")
    wb.save(dateipfad)

def frage(text = ""):
    print_(text + " Zur Fortsetzung Taste Drücken.")
    eingabe = input()
#____________________________________ PROGRAMMSTART ___________________________________________


updatedList = updateList(lizenzpool_neu,lizenzpool_alt,True)     # Alte und Neue Liste integrieren
listToExcel("ergebnis",updatedList)                         # liste in Excel speichern
loescheExcelSpalte('ergebnis.xlsx',1)                       # Dataframe_Spalte löschen (Kovertierungsproblem)
autoSpaltenBreite('ergebnis.xlsx')                          # autoSpaltenBreite
mergeCells('ergebnis.xlsx','typ')                           # Zellen "typ" mergen
mergeCells('ergebnis.xlsx','leistungsmerkmal')              # Zellen "leistungsmerkmal" mergen
ergebnisKopfzeileDesign()
zeilenHoehe()
